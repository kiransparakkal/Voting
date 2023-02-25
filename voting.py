
from openpyxl import Workbook
from collections import Counter


"""
GeneratePreferences(values) inputs a set of numerical values that the agents 
have for the different alternatives and outputs a preference profile.

"""
def generatePreferences(values):
    
 
    preferences ={}#dictionary to be returned 
    rownumber = 1# initialise rownumber
   
    for row in values.values: #iterate through each rows
       sample_list1=[]
       sample_list2=[]
       dictvalue =[]
       sample_list1=list(row)#values of each row is copied onto a list
    
       """
        Below for loop iterates over first list and save the index and values of samplelist1
        onto samplelist2 as tuple pairs (x,y) onto list(x=alterntive,y=value of alternative)
       """
       for x,y in enumerate(sample_list1):
         sample_list2.append((x+1,y))
       
       #samplelist2 is first sorted and then reversed
       sample_list2.sort(key=lambda x:(x[1],x[0]))
       sample_list2.reverse()

       #Copies only the alternative no corresponding to each value after sorting onto dictvalue list
       for item in sample_list2:
         dictvalue.append(item[0])
       #Inputs a set of numerical values that the agents have for the different alternatives onto preferences{}
       preferences[rownumber]= dictvalue

       rownumber+=1
    return preferences
  
    
"""
The function inputs a preference profile and 
returns a winner according to the Dictatorship rule
"""
def dictatorship(preferenceProfile, agent):
    if (agent in preferenceProfile.keys() and type(agent) == int):
        return preferenceProfile[agent][0]
    else:
        print("Invalid agent")
        return False
"""
The function handles tie situation between possible winners
and returns the appropriate winner according to 'tieBreak" input
"""
def tieBreaker(tieBreak,preferenceProfile,value_return):
     
     if tieBreak == 'max':
        return max(value_return)
    
     if tieBreak == 'min':
        return min(value_return)
    
     if (type(tieBreak) == int and tieBreak >0 and tieBreak<= max(preferenceProfile.keys())):
        for x in preferenceProfile[tieBreak]:
          if x in (value_return):
            return x
     else:
       print("Invalid agent number")
       return False

"""
The function inputs a preference profile and 
returns a winner according to the Plurality rule
"""
def plurality(preferenceProfile, tieBreak):

   
    # get every agents first preferences
    first_preferences = []
    value_return = []#sample list of possible winners
     
    #Stores the first preference of each agent onto a list 
    for key in preferenceProfile.keys():
        first_preferences.append(preferenceProfile[key][0])

    count = Counter(first_preferences)# creates a dictionary with elements of list as key and its no of occurances as value.

    max_occurance = max(count.values())#gets the max count value

    for key in count.keys():
        if count[key] == max_occurance:
            value_return.append(key)#possbile winners are stored onto value_return list
    
    #when there is only one possible winner
    if len(value_return) == 1:
        return value_return[0]
    #when there are more than one possible winner
    return(tieBreaker(tieBreak,preferenceProfile,value_return))
"""
The function inputs a preference profile and 
returns a winner according to the Veto rule
"""         
    
def veto(preferenceProfile, tieBreak):

    value_return = []

    # get everyones preferences except last preference
    veto_preferences = []
    for l in preferenceProfile.keys():
        veto_preferences = veto_preferences + preferenceProfile[l][:-1]
    
    count = Counter(veto_preferences)# creates a dictionary with each alternative as key and thier corresponding points as values.

    max_occurance = max(count.values())#gets the max count value

    #Select possible winners based on those having most number of points 
    for key in count.keys():
        if count[key] == max_occurance:
            value_return.append(key)#possbile winners are stored onto value_return list
    
    #when there is only one possible winner
    if len(value_return) == 1:
        return value_return[0]
    
    #when there are more than one possible winner
    return(tieBreaker(tieBreak,preferenceProfile,value_return))

"""
The function inputs a preference profile and 
returns a winner according to the Scoring rule
"""  

def scoringRule(preferenceProfile, scoreVector, tieBreak):

    score = {} #Dictionary to store each alternative with their corresponding total score 
    number_of_alternatives = len(list(preferenceProfile.values())[0])

    #Returns False when the len(scoringVector) is less than no of alternatives
    if (len(scoreVector)< number_of_alternatives):
      print("Incorrect input")
      return False
    
    scoreVector.sort(reverse=True)
    for i in range(1, number_of_alternatives + 1):
        score[i] = 0 #Initialises score of all alternatives as 0
    #Initialises each alternatives with their corresponding total score
    for x in preferenceProfile.keys():
        p=0
        for value in preferenceProfile[x]:
            score[value] += scoreVector[p]
            p+= 1
        
    max_score = max(score.values())#Gets max score assigned to any alternatives in dictionary
    value_return = []
    #Possible winning alternatives are selected based upon those having max total score 
    for key in score.keys():
        if score[key] == max_score:
            value_return.append(key)#possbile winners are stored onto value_return list
    
    #when there is only one possible winner
    if len(value_return) == 1:
        return value_return[0]
    
    #when there are more than one possible winner
    return(tieBreaker(tieBreak,preferenceProfile,value_return))

"""
The function inputs a preference profile and 
returns a winner according to the rangeVoting rule
"""  
   
def rangeVoting (values, tieBreak):
   score = {}#Dictionary to store each alternative with their corresponding total score 
   
   number_of_alternatives =values.max_column#Gets the total no of alternatives
   number_of_agents = values.max_row#Gets the totals no of agents
   
   for i in range(1, number_of_alternatives + 1):
        score[i] = 0#Initialises score of all alternatives as 0
   
   """
   Assigns each alternative with their corresponding sum of valuations in the form of 'score' dictionary
   where 'key' denotes altenative and 'value' denotes sum of its corresponding valuations
   """ 
   for row in values.iter_rows():
          x=1
          for cell in row:
            score[x] += cell.value
            x+=1
   max_score = max(score.values())#maximum sum of valuations
   agent_dict={}
   
   #Below code is used to get the specified agent preference profile
   if(type(tieBreak)==int):
    sample_dict={}
    agent_dict_values=[]
    x=1
    #Iternates through values of row corrensponding to agent no = 'tieBreak'
    for row in values.iter_rows(min_row=tieBreak, max_row=tieBreak):
      for cell in row:
        sample_dict[x]=cell.value#saves the alternatives with their values onto dict
        x+=1
    for i in sorted(sample_dict, key=sample_dict.get):
        agent_dict_values.append(i)
    agent_dict_values.reverse()
    agent_dict[tieBreak]=agent_dict_values

   value_return = []

   #Possible winning alternatives are selected based upon those having max sum of valuations
   for key in score.keys():
        if score[key] == max_score:
            value_return.append(key)
   #when there is only one possible winner
   if len(value_return) == 1:
        return value_return[0]
   #when there are more than one possible winner
   return(tieBreaker(tieBreak,agent_dict,value_return))
   
"""
The function inputs a preference profile and 
returns a winner according to the borda rule
"""  
def borda(preferenceProfile, tieBreak):

    score = {}#Dictionary to store each alternative with their corresponding total score
   

    number_of_alternatives = len(list(preferenceProfile.values())[0])
    for i in range(1, number_of_alternatives + 1):
        score[i] = 0 #Initialises score of all alternatives as 0
    
    #Assigns a score value of m-j to the alternative ranked at position j,m is total no of alternatives
    for x in preferenceProfile.keys():

        sample_value = 1
         
        for value in preferenceProfile[x]:
            score[value] += number_of_alternatives - sample_value
            sample_value += 1
        
    max_score = max(score.values())

    value_return = []

    #Possible winning alternatives are selected based upon those having max score value
    for key in score.keys():
        if score[key] == max_score:
            value_return.append(key)

    #when there is only one possible winner
    if len(value_return) == 1:
        return value_return[0]
    
    #when there are more than one possible winner
    return(tieBreaker(tieBreak,preferenceProfile,value_return))
"""
The function inputs a preference profile and 
returns a winner according to the harmonic rule
""" 

def harmonic(preferenceProfile, tieBreak):

    score = {}#Dictionary to store each alternative with their corresponding total score

    number_of_alternatives = len(list(preferenceProfile.values())[0])
    for i in range(1, number_of_alternatives + 1):
        score[i] = 0#Initialises score of all alternatives as 0

    #Assigns a score value of 1/j to the alternative ranked at position j
    for key in preferenceProfile.keys():

        sample_value = 1


        for value in preferenceProfile[key]:
            score[value] += 1/sample_value
            sample_value += 1
    
    max_score = max(score.values())

    value_return = []

    #Possible winning alternatives are selected based upon those having max score value
    for key in score.keys():
        if score[key] == max_score:
            value_return.append(key)

    #when there is only one possible winner
    if len(value_return) == 1:
        return value_return[0]
    
    #when there are more than one possible winner
    return(tieBreaker(tieBreak,preferenceProfile,value_return))
"""
The function inputs a preference profile and 
returns a winner according to the STV rule
""" 

def STV (preferenceProfile, tieBreak):
    score = {}#Dictionary to store each alternative with their corresponding total score
    value_return = []#list that store possible winners

    
    number_of_alternatives = len(list(preferenceProfile.values())[0])#total no of alternatives
    
    for i in range(1, number_of_alternatives + 1):
        score[i] = 0#Initialises score of all alternatives as 0

    """
    Assigns score with 'key' as first preference alternative of each agent and
    'values' as their frequency of occurance in position 1
    """     
    for key in preferenceProfile.keys():
        score[preferenceProfile[key][0]]+=1
    
    #Loop inorder to conduct different rounds to select winners as per STV
    while True:
      max_score=max(score.values())#max value in score dictionary
      min_score=min(score.values())#min value in score dictionary
      
      value_to_remove=[]#list that stores alternatives to be removed

      for item in score.keys():
        if(score[item] == min_score and max_score != min_score):
            value_to_remove.append(item)
        elif(score[item] == min_score and min_score == max_score):
            value_return.append(item)

      if(len(value_return)>0):
        break
      else:
       #Removes the alternatives to be removed from score dictionary 
       for x in value_to_remove:
        score.pop(x)


    #when there is only one possible winner
    if len(value_return) == 1:
        return value_return[0]
    
    #when there are more than one possible winner
    return(tieBreaker(tieBreak,preferenceProfile,value_return))
    


           
       


